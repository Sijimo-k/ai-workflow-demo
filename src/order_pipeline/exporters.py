"""成果物の出力（統合CSV / 出荷リストExcel / メール文面ファイル群）。"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .mail import mail_filename, render_mail
from .models import Order

_UNIFIED_HEADER = [
    "モール",
    "注文番号",
    "注文日時",
    "顧客名",
    "メールアドレス",
    "商品名",
    "単価",
    "数量",
    "金額",
    "郵便番号",
    "住所",
    "入金状態",
    "発送状態",
]


def export_unified_csv(orders: list[Order], path: Path) -> Path:
    """統合済み全受注をCSV出力（UTF-8 BOM付き = Excelで文字化けしない）。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(_UNIFIED_HEADER)
        for o in orders:
            writer.writerow(
                [
                    o.source.value,
                    o.order_id,
                    o.order_date.strftime("%Y-%m-%d %H:%M:%S"),
                    o.customer_name,
                    o.customer_email,
                    o.product_name,
                    o.unit_price,
                    o.quantity,
                    o.amount,
                    o.postal_code,
                    o.address,
                    o.payment_status.value,
                    o.fulfillment_status.value,
                ]
            )
    return path


_SHIPPING_HEADER = [
    "注文番号",
    "モール",
    "注文日",
    "顧客名",
    "商品名",
    "数量",
    "金額",
    "郵便番号",
    "住所",
]


def export_shipping_list_excel(targets: list[Order], path: Path) -> Path:
    """発送対象の出荷リストをExcel出力。ヘッダ装飾・列幅・合計行つき。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "出荷リスト"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(_SHIPPING_HEADER)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for o in targets:
        ws.append(
            [
                o.order_id,
                o.source.value,
                o.order_date.strftime("%Y-%m-%d"),
                o.customer_name,
                o.product_name,
                o.quantity,
                o.amount,
                o.postal_code,
                o.address,
            ]
        )

    total_row = ["", "", "", "", "合計", sum(o.quantity for o in targets), sum(o.amount for o in targets), "", ""]
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    _autofit_columns(ws)
    wb.save(path)
    return path


def _autofit_columns(ws) -> None:
    """全角を2幅として概算で列幅を合わせる。"""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 0
        for cell in ws[letter]:
            text = "" if cell.value is None else str(cell.value)
            cell_width = sum(2 if ord(ch) > 0x2E7F else 1 for ch in text)
            width = max(width, cell_width)
        ws.column_dimensions[letter].width = min(width + 2, 50)


def export_mails(targets: list[Order], outdir: Path) -> list[Path]:
    """発送対象ごとにメール文面を .txt 出力。書き出したパス一覧を返す。"""
    outdir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for o in targets:
        path = outdir / mail_filename(o)
        path.write_text(render_mail(o), encoding="utf-8")
        written.append(path)
    return written
